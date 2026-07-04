from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl

from src.models import Task
from src.validators import validate_required_columns


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def available_sheets(input_path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_tasks(
    input_path: Path,
    sheet: str = "all",
    domain: str | None = None,
    limit: int | None = None,
) -> list[Task]:
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    tasks: list[Task] = []
    try:
        target_sheets = _resolve_target_sheets(workbook.sheetnames, sheet)
        for sheet_name in target_sheets:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if header_row is None:
                continue

            headers = [_cell_str(h) for h in header_row]
            validate_required_columns(headers)
            idx = {column: position for position, column in enumerate(headers)}

            for row in rows:
                task = _row_to_task(sheet_name, row, headers, idx)
                if domain and task.global_domain.lower() != domain.lower():
                    continue
                tasks.append(task)

                if limit and len(tasks) >= limit:
                    return tasks
    finally:
        workbook.close()

    return tasks


def _resolve_target_sheets(all_sheet_names: Iterable[str], requested_sheet: str) -> list[str]:
    all_sheets = list(all_sheet_names)
    if requested_sheet.lower() == "all":
        return all_sheets

    normalized_map = {name.lower(): name for name in all_sheets}
    try:
        return [normalized_map[requested_sheet.lower()]]
    except KeyError as exc:
        available = ", ".join(all_sheets)
        raise ValueError(
            f"Sheet '{requested_sheet}' not found. Available sheets: {available}"
        ) from exc


def _row_to_task(
    sheet_name: str,
    row: tuple[object, ...],
    headers: list[str],
    idx: dict[str, int],
) -> Task:
    metadata = {column: _cell_str(row[pos]) for column, pos in idx.items()}

    return Task(
        sheet_name=sheet_name,
        batch=_cell_str(row[idx["batch"]]),
        task_number=_cell_str(row[idx["task_number"]]),
        task_name=_cell_str(row[idx["task name"]]),
        global_domain=_cell_str(row[idx["global domain"]]),
        task_type=_cell_str(row[idx["task_type"]]),
        manual_status=_cell_str(row[idx["manual_status"]]),
        final_prompt_text=_cell_str(row[idx["final_prompt_text"]]),
        final_prompt_text_ru=_cell_str(row[idx["final_prompt_text_ru"]]),
        fixed_criteria_0_2_each=_cell_str(row[idx["fixed_criteria_0_2_each"]]),
        fixed_criteria_0_2_each_ru=_cell_str(row[idx["fixed_criteria_0_2_each_ru"]]),
        rubric_id=_cell_str(row[idx["rubric_id"]]),
        gold_set_id=_cell_str(row[idx["gold_set_id"]]),
        correct_items=_cell_str(row[idx["correct_items"]]),
        acceptable_variants=_cell_str(row[idx["acceptable_variants"]]),
        optional_distractors=_cell_str(row[idx["optional_distractors"]]),
        explicit_exclusion_rules=_cell_str(row[idx["explicit_exclusion_rules"]]),
        boundary_notes=_cell_str(row[idx["boundary_notes"]]),
        metadata=metadata,
    )

