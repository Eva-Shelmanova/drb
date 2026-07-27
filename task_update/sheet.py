"""Read/parse/write helpers for the Core sheet of final.xlsx."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

SHEET = "Core"

# The five rubric dimensions, keyed by sheet column name.
DIMENSIONS = ["coverage", "accuracy", "reasoning", "use_of_evidence", "clarity_and_structure"]

# The boundary sentence exists in two forms. The sheet originally labelled it
# ("Methodological trap: distinguish ..."), which tells the model under test
# that a trick is present. The neutral form drops the label and just states the
# distinctions. Both are recognised so the parsers work before and after
# neutralisation.
EN_TRAP_LABELLED_RE = re.compile(
    r"(?:Methodological trap|Boundary trap)\s*:\s*(?P<body>[Dd]istinguish\b.*?)"
    r"(?=\s+Return\s*:|\s*$)",
    re.DOTALL,
)
EN_TRAP_NEUTRAL_RE = re.compile(
    r"(?P<body>Distinguish\b.*?)(?=\s+Return\s*:|\s*$)",
    re.DOTALL,
)
# Russian prompts use several return-clause spellings.
RU_RETURN_RE = re.compile(
    r"(?:Результат|Итог|Итоги|Возврат|Возвращаемое значение|Вывод|Верните)\s*:"
)
RU_TRAP_MARKER_RE = re.compile(r"(?:ловушк\w*|ошибк\w*)\s*:", re.IGNORECASE)
RU_NEUTRAL_RE = re.compile(r"Различа\w*\b")

BOLD_RE = re.compile(r"\*\*([^*]+)\*\*")

# A rubric header is a bold span immediately followed by a "- 2:" level marker.
# Anchoring on the level marker is what separates headers from bolded gold
# values, which matters because the Russian cells bold both and carry no
# newlines at all. Verified to yield exactly 5 headers on all 161 Core rows in
# both languages.
# The whitespace after the header stays outside the match so that a parse /
# compose round trip reproduces the cell byte for byte.
RUBRIC_HDR_RE = re.compile(r"\*\*(?P<header>[^*]{3,60}?)\*\*(?=\s*-\s*2\s*:)")


class Row(dict):
    """A sheet row plus its 1-based worksheet row index."""

    def __init__(self, values: dict[str, Any], row_index: int) -> None:
        super().__init__(values)
        self.row_index = row_index

    @property
    def row_id(self) -> str:
        return (
            f"{SHEET}-batch{self.get('batch')}"
            f"-task{self.get('task_number')}-{self.get('task name')}"
        )


def load_rows(xlsx: Path) -> tuple[list[str], list[Row]]:
    """Return (header, rows) for the Core sheet."""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]

    rows: list[Row] = []
    for i, raw in enumerate(it, start=2):
        if not any(c not in (None, "") for c in raw):
            continue
        values = {h: ("" if v is None else str(v)) for h, v in zip(header, raw)}
        rows.append(Row(values, row_index=i))
    wb.close()
    return header, rows


# --------------------------------------------------------------------------
# trap parsing
# --------------------------------------------------------------------------

def parse_en_trap(prompt: str) -> tuple[str, str] | None:
    """Return (full_boundary_sentence, body) or None.

    The body always starts at "distinguish", so it is identical in both the
    labelled and neutral forms; only the returned sentence differs.
    """
    m = EN_TRAP_LABELLED_RE.search(prompt) or EN_TRAP_NEUTRAL_RE.search(prompt)
    if not m:
        return None
    return m.group(0).strip(), m.group("body").strip()


def parse_ru_trap(prompt: str) -> tuple[str, str] | None:
    """Locate the Russian boundary sentence, then bound it by sentence edges."""
    m = RU_TRAP_MARKER_RE.search(prompt)
    neutral = m is None
    if neutral:
        m = RU_NEUTRAL_RE.search(prompt)
    if not m:
        return None

    # Sentence start: after the previous period, else start of text.
    start = prompt.rfind(". ", 0, m.start())
    start = 0 if start == -1 else start + 2

    # Sentence end: before the return clause if present, else the next period.
    ret = RU_RETURN_RE.search(prompt, m.end())
    if ret:
        end = ret.start()
    else:
        dot = prompt.find(".", m.end())
        end = len(prompt) if dot == -1 else dot + 1

    sentence = prompt[start:end].strip()
    # In the labelled form the body starts after the marker; in the neutral form
    # the sentence is the body.
    body = (sentence if neutral else prompt[m.end():end]).strip()
    if not sentence:
        return None
    if not neutral and not re.search(r"ловушк|ошибк", sentence, re.IGNORECASE):
        return None
    return sentence, body


def neutralise_en(sentence: str) -> str:
    """Drop the 'Methodological trap:' label, keeping the distinctions."""
    m = EN_TRAP_LABELLED_RE.search(sentence)
    if not m:
        return sentence.strip()
    body = m.group("body").strip()
    return body[:1].upper() + body[1:]


def neutralise_ru(sentence: str) -> str:
    """Drop the Russian trap label, keeping the distinctions."""
    m = RU_TRAP_MARKER_RE.search(sentence)
    if not m:
        return sentence.strip()
    body = sentence[m.end():].strip()
    return body[:1].upper() + body[1:]


def count_distinctions(body: str) -> int:
    """Count the distinctions enumerated in a trap body."""
    text = re.sub(r"^\s*distinguish(?:\s+between)?\s*", "", body, flags=re.IGNORECASE)
    text = re.sub(r"^\s*различ\w*\s*", "", text, flags=re.IGNORECASE)
    text = text.rstrip(" .")
    parts = re.split(r";|,|\band\b|\bи\b", text)
    return len([p for p in parts if len(p.strip()) > 2])


def splice_trap(prompt: str, old_sentence: str, new_sentence: str) -> str:
    """Replace the trap sentence in place, leaving everything else byte-identical."""
    if old_sentence not in prompt:
        raise ValueError("old trap sentence not found in prompt")
    return prompt.replace(old_sentence, new_sentence, 1)


# --------------------------------------------------------------------------
# rubric composition
# --------------------------------------------------------------------------

def flatten_ws(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


class RubricBlock:
    """One rubric dimension, with the exact whitespace that surrounded it."""

    __slots__ = ("header", "sep", "body", "tail")

    def __init__(self, header: str, sep: str, body: str, tail: str) -> None:
        self.header = header
        self.sep = sep
        self.body = body
        self.tail = tail

    def render(self, body: str | None = None) -> str:
        return f"**{self.header}**{self.sep}{self.body if body is None else body}{self.tail}"


def parse_rubric(fixed_criteria: str) -> tuple[str, list[RubricBlock]]:
    """Split a rubric cell into (prefix, blocks), preserving all whitespace."""
    matches = list(RUBRIC_HDR_RE.finditer(fixed_criteria))
    if not matches:
        return fixed_criteria, []
    prefix = fixed_criteria[: matches[0].start()]
    blocks = []
    for i, m in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(fixed_criteria)
        raw = fixed_criteria[m.end():end]
        sep = raw[: len(raw) - len(raw.lstrip())]
        tail = raw[len(raw.rstrip()):]
        blocks.append(RubricBlock(m.group("header").strip(), sep, raw.strip(), tail))
    return prefix, blocks


def rubric_bodies(fixed_criteria: str) -> dict[str, str]:
    """Map dimension key -> body, matched positionally against DIMENSIONS."""
    _, blocks = parse_rubric(fixed_criteria)
    if len(blocks) != len(DIMENSIONS):
        raise ValueError(f"expected {len(DIMENSIONS)} rubric blocks, found {len(blocks)}")
    return {DIMENSIONS[i]: b.body for i, b in enumerate(blocks)}


def compose_rubric(fixed_criteria: str, new_bodies: dict[str, str]) -> str:
    """Rebuild the rubric cell, substituting the given dimension bodies.

    Headers, their order, their casing and the surrounding whitespace all come
    from the existing cell, so `compose_rubric(text, {}) == text` exactly.
    Bodies are matched to dimensions by position, because the Russian header
    wording varies across rows while the order does not.
    """
    prefix, blocks = parse_rubric(fixed_criteria)
    if len(blocks) != len(DIMENSIONS):
        raise ValueError(f"expected {len(DIMENSIONS)} rubric blocks, found {len(blocks)}")

    # Russian cells are a single flat line; English cells use newlines.
    flat = "\n" not in fixed_criteria
    out = [prefix]
    for i, block in enumerate(blocks):
        new = new_bodies.get(DIMENSIONS[i])
        if new is None:
            out.append(block.render())
        else:
            out.append(block.render(flatten_ws(new) if flat else str(new).strip()))
    return "".join(out)


def gold_spans(row: Row, ru: bool = False) -> list[str]:
    """Bolded gold values, used to detect answer leakage into distractors."""
    suffix = "_ru" if ru else ""
    spans: list[str] = []
    for col in (f"accuracy{suffix}", f"key_facts_from_source_bundle{suffix}"):
        spans.extend(s.strip() for s in BOLD_RE.findall(row.get(col, "")))
    seen, uniq = set(), []
    for s in spans:
        k = re.sub(r"\s+", " ", s.lower())
        if k and k not in seen:
            seen.add(k)
            uniq.append(s)
    return uniq


def bullets(items: list[str]) -> str:
    return "\n".join(f"- {i.strip().lstrip('-').strip()}" for i in items if i and i.strip())


def append_bullets(existing: str, items: list[str]) -> str:
    new = bullets(items)
    if not new:
        return existing
    if not existing.strip():
        return new
    return existing.rstrip() + "\n" + new


# --------------------------------------------------------------------------
# write-back
# --------------------------------------------------------------------------

def write_updates(src: Path, dst: Path, updates: dict[int, dict[str, str]]) -> int:
    """Copy the workbook to dst with the given {row_index: {column: value}} edits."""
    wb = load_workbook(src)
    ws = wb[SHEET]
    header = [
        str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))
    ]
    col_of = {h: i + 1 for i, h in enumerate(header)}

    written = 0
    for row_index, fields in updates.items():
        for col, value in fields.items():
            if col not in col_of:
                raise KeyError(f"unknown column: {col}")
            ws.cell(row=row_index, column=col_of[col], value=value)
        written += 1

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()
    return written


def iter_columns(header: list[str]) -> Iterator[str]:
    yield from header
